To accomplish this task, we will create a data-driven testing framework using Pylest (assuming this is a typo for PyTest), Page Object Model (POM), and Excel for storing test data and results. We will use Selenium WebDriver for browser automation and `openpyxl` for interacting with Excel files.

### Step-by-Step Guide:

1. Set Up Project Environment:
   - Install required packages:
     ```bash
     pip install selenium openpyxl pytest
     ```

2. Create Excel File:
   - Create an Excel file (`test_data.xlsx`) with columns: `Test ID`, `Username`, `Password`, `Date`, `Time of Test`, `Name of Tester`, `Test Result`.

3. Page Object Model (POM):
   - Create a page object model for the login page.

4. Write Test Script:
   - Read data from the Excel file.
   - Perform login actions using the data.
   - Write results back to the Excel file.

### Directory Structure:
```
.
├── pages
│   └── login_page.py
├── tests
│   └── test_login.py
├── utils
│   └── excel_utils.py
└── test_data.xlsx
```

### 1. Create Excel File (`test_data.xlsx`):
Fill this file with test data as per your requirement.

### 2. Page Object Model (POM):
pages/login_page.py:
```python
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

class LoginPage:
    def __init__(self, driver):
        self.driver = driver
        self.username_input = (By.ID, 'username')  # Adjust selector as needed
        self.password_input = (By.ID, 'password')  # Adjust selector as needed
        self.login_button = (By.ID, 'loginButton')  # Adjust selector as needed

    def enter_username(self, username):
        WebDriverWait(self.driver, 10).until(
            EC.visibility_of_element_located(self.username_input)).send_keys(username)

    def enter_password(self, password):
        WebDriverWait(self.driver, 10).until(
            EC.visibility_of_element_located(self.password_input)).send_keys(password)

    def click_login(self):
        WebDriverWait(self.driver, 10).until(
            EC.element_to_be_clickable(self.login_button)).click()

    def is_login_successful(self):
        # Adjust condition to determine successful login
        return WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.ID, 'successfulLoginElement')))
```

### 3. Utility for Excel Operations:
utils/excel_utils.py:
```python
import openpyxl
from datetime import datetime

class ExcelUtils:
    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = openpyxl.load_workbook(file_path)
        self.sheet = self.workbook.active

    def get_test_data(self):
        test_data = []
        for row in range(2, self.sheet.max_row + 1):
            data = {
                "test_id": self.sheet.cell(row=row, column=1).value,
                "username": self.sheet.cell(row=row, column=2).value,
                "password": self.sheet.cell(row=row, column=3).value,
                "tester": self.sheet.cell(row=row, column=6).value
            }
            test_data.append(data)
        return test_data

    def write_result(self, test_id, result):
        for row in range(2, self.sheet.max_row + 1):
            if self.sheet.cell(row=row, column=1).value == test_id:
                self.sheet.cell(row=row, column=7, value=result)
                self.sheet.cell(row=row, column=4, value=datetime.now().date())
                self.sheet.cell(row=row, column=5, value=datetime.now().time())
                break
        self.workbook.save(self.file_path)
```

### 4. Test Script:
tests/test_login.py:
```python
import pytest
from selenium import webdriver
from pages.login_page import LoginPage
from utils.excel_utils import ExcelUtils

@pytest.fixture(scope="module")
def driver():
    driver = webdriver.Chrome()
    driver.get("https://opensouro-demo.grangchrmliva.com/b/index.pha/auth/login")
    yield driver
    driver.quit()

def test_login(driver):
    excel_utils = ExcelUtils('test_data.xlsx')
    test_data = excel_utils.get_test_data()
    login_page = LoginPage(driver)
    
    for data in test_data:
        login_page.enter_username(data['username'])
        login_page.enter_password(data['password'])
        login_page.click_login()

        try:
            assert login_page.is_login_successful()
            excel_utils.write_result(data['test_id'], 'Pass')
        except Exception as e:
            excel_utils.write_result(data['test_id'], 'Fail')
        driver.get("https://opensouro-demo.grangchrmliva.com/b/index.pha/auth/login")
```

### Running the Tests:
- Run the tests using PyTest:
  ```bash
  pytest tests/test_login.py
  ```

### Notes:
- Adjust the element locators in `login_page.py` according to the actual IDs or selectors used in the web page.
- Make sure the Excel file `test_data.xlsx` is properly formatted with valid test data.
- Ensure the WebDriver executable (e.g., chromedriver) is in your PATH or specify its location when initializing the WebDriver.

This setup should help you perform data-driven testing for the login functionality of your portal using Python, Selenium, and PyTest.
